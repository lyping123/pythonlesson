# from openpyxl import Workbook

# wordbook=Workbook()
# sheet=wordbook.active
# sheet.title="MySheet"

# sheet['A1'] = 87  
# sheet['A2'] = "Devansh"  
# sheet['A3'] = 100.80  
# sheet['A4'] = 200

# wordbook.save("demo.xlsx")


# from openpyxl import load_workbook  
# filepath="demo.xlsx"
# wb = load_workbook(filepath)  
  
# sheet = wb.active  
# sheet['A1'] = 'Devansh Sharma'  
  
# sheet.cell(row=2, column=2).value = 5  
# wb.save("demo.xlsx") 

# from openpyxl import Workbook  
  
# wb = Workbook()  
# sheet = wb.active  

# data = (  
#     (11, 48, 50,100),  
#     (81, 30, 82),  
#     (20, 51, 72),  
#     (21, 14, 60),  
#     (28, 41, 49),  
#     (74, 65, 53),  
#     ("Peter", 'Andrew',45.63)  
# )  

# for i in data:    
#     sheet.append(i)
#     # sheet.cell(row=sheet.max_row+1, column=1, value=i[0])
      
# wb.save('demo.xlsx')  


#read
# import openpyxl  
  
# wb = openpyxl.load_workbook('demo.xlsx')  
  
# sheet = wb.active  
  
# # x1 = sheet['A1']  
# # x2 = sheet['A2']  

# cells=sheet["A1:B7"]
# #using cell() function  
# # x3 = sheet.cell(row=3, column=1)  
  
# # print("The first cell value:",x1.value)  
# # print("The second cell value:",x2.value)  
# # print("The third cell value:",x3.value) 

# for i1,i2 in cells:
    
#     print("{0:10} {1:10}".format(i1.value,i2.value))  



from openpyxl import Workbook  
  
wb = Workbook()  
sheet = wb.active  
  
rows = (  
    (90, 46, 48, 44),  
    (81, 30, 32, 16),  
    (23, 95, 87,27),  
    (65, 12, 89, 53),  
    (42, 81, 40, 44),  
    (34, 51, 76, 42)  
)  
  
for row in rows:  
    sheet.append(row)  
  
for row in sheet.iter_rows(min_row=2, min_col=2, max_row=4, max_col=4):  
    for cell in row:  
        print(cell.value, end=" ")  
    print()  
  
wb.save('iter_rows.xlsx')  

